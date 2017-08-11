# encoding:utf-8
import xlwt
import re
from xlsxwriter import *
import openpyxl
import glob

filelocation = '/Users/chenzhaopeng/PycharmProjects/txt2xls/final_data/'
read_file_form = "txt"
fileform = "xlsx"
filedestination = "/Users/chenzhaopeng/PycharmProjects/txt2xls/"
final_file = "merge_all_information"

#读取txt文件夹里有多少文件
filearray = []
for filename in glob.glob(filelocation + "*." + read_file_form):
    filearray.append(filename)
    print(filename)
print("在文件夹内有 %d 个txt文档"% len(filearray))

#添加表头
biaotou = ['股票代码', '股票简称', '会计期间', '数据期间', '项目名称', '项目金额', '计价货币', '说明','公司名称']
final_excel = openpyxl.Workbook()
final_excel_sheet = final_excel.active
final_excel_sheet.append(biaotou)

final_excel_row = 1

#没有合计的公司的数量
no_all_company_num = 0
#没有表格的公司的数量
no_table_num = 0

#将没有表格的公司写入另外的TXT文件
no_table_file = open("no_table_company.txt",'w+')
#将没有任何patten的公司写入另外的TXT文件
no_double_num_file = open("no_double_num_file.txt",'w+')
#有多个number连在一起的
lots_of_number_file = open("lost_of_number_file.txt","w+")

# no_all_file = open("no_all_company.txt",'w+')

#比较两个数字的大小,取最大的一个,供10年之前需要比较好几个销售数量的情况使用
def is_bigger(str1,str2):
    if len(str1) < len(str2):
        return -1
    if len(str1) > len(str2):
        return 1
    new_str1=str1.replace(',', '')
    new_str1=new_str1.replace('，', '')
    new_str1=new_str1.replace('.', '')
    int1 = int(new_str1)
    new_str2=str2.replace(',', '')
    new_str2=new_str2.replace('，', '')
    new_str2=new_str2.replace('.', '')
    int2 = int(new_str2)
    return int1-int2

#逐个读取文件
for txtname in filearray:
    txtfile = open(txtname,'r')
    line = txtfile.read()

    #先按照——————划分出不同公司
    company_list = line.split('---------------------------------')
    # print(company)
    # for i in range(100):
    for i in range(len(company_list)-1):
    # for company in company_list:
        data_time = None
        pro_name = None
        pro_num = None
        company_name = None
        other = None
        money = 'CNY'

        #each_line 公司里面的每个项目,空格分开
        each_line = re.findall(r'[^\s]+',company_list[i])
        # print(each_line)
        code = each_line[0]
        short_name = each_line[1]
        count_time = each_line[2]
        company_name = each_line[3]
        #如果没有第五个,没有表格,输出新的TXT
        if len(each_line) <= 5:
            no_table_file.write(company_list[i])
            no_table_num += 1
            continue

        #如果公司没有表格,直接continue,并且输出到新的txt
        if each_line[4] == 'None':
            # print(code + "没有表格")
            no_table_file.write(company_list[i])
            no_table_num += 1
            continue

        # 只有一行销售费用，没有具体项目,适用于10年之前的,仅含有销售费用
        #但销售费用有多个,取其中的最大的两个,左边的是今年的,右边的是去年的
        fine_sell = re.findall(r"^\s*销",each_line[4])
        if len(fine_sell) != 0:
            box = ''
            for back in range(5, min(10, len(each_line))):
                is_nag = re.findall(r'-\d|%|％|%|[\u4e00-\u9fa5]', each_line[back])
                if len(is_nag) > 0:
                    continue
                is_int = re.findall(r'\d,|\.\d\d', each_line[back])
                if len(is_int) > 0:
                    box = box + ' ' + str(each_line[back])
            each_number = re.findall(r'[^\s]+', box)
            #去掉 括号包着数字的情况
            for temp in range(len(each_number)):
                each_number[temp] = each_number[temp].replace('(', '')
                each_number[temp] = each_number[temp].replace(')', '')
            if len(each_number) == 1:
                row_now = [code, short_name, count_time, 1, '销售费用', each_number[0], money, other, company_name]
                final_excel_row += 1
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
            elif len(each_number) == 2:
                row_now = [code, short_name, count_time, 1, '销售费用', each_number[0], money, other, company_name]
                final_excel_row += 1
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
                row_now = [code, short_name, count_time, 0, '销售费用', each_number[1], money, other, company_name]
                final_excel_row += 1
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
            elif len(each_number) == 3:
                min_local = 0
                for problem in range(len(each_number)):
                    if is_bigger(each_number[min_local], each_number[problem]) > 0:
                        min_local = problem
                del each_number[min_local]
                row_now = [code, short_name, count_time, 1, '销售费用', each_number[0], money, other, company_name]
                final_excel_row += 1
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
                row_now = [code, short_name, count_time, 0, '销售费用', each_number[1], money, other, company_name]
                final_excel_row += 1
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
            elif len(each_number) == 4:
                min_local = 0
                for problem in range(1, len(each_number)):
                    if is_bigger(each_number[min_local], each_number[problem]) > 0:
                        min_local = problem
                del each_number[min_local]
                min_local = 0
                for problem in range(1, len(each_number)):
                    if is_bigger(each_number[min_local], each_number[problem]) > 0:
                        min_local = problem
                del each_number[min_local]
                row_now = [code, short_name, count_time, 1, '销售费用', each_number[0], money, other, company_name]
                final_excel_row += 1
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
                row_now = [code, short_name, count_time, 0, '销售费用', each_number[1], money, other, company_name]
                final_excel_row += 1
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
            elif len(each_number) == 5:
                min_local = 0
                for problem in range(1, len(each_number)):
                    if is_bigger(each_number[min_local], each_number[problem]) > 0:
                        min_local = problem
                del each_number[min_local]
                min_local = 0
                for problem in range(1, len(each_number)):
                    if is_bigger(each_number[min_local], each_number[problem]) > 0:
                        min_local = problem
                del each_number[min_local]
                min_local = 0
                for problem in range(1, len(each_number)):
                    if is_bigger(each_number[min_local], each_number[problem]) > 0:
                        min_local = problem
                del each_number[min_local]
                row_now = [code, short_name, count_time, 1, '销售费用', each_number[0], money, other, company_name]
                final_excel_row += 1
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
                row_now = [code, short_name, count_time, 0, '销售费用', each_number[1], money, other, company_name]
                final_excel_row += 1
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
            continue

        # lot_of_number_condition_list = None
        # lot_of_number_condition_list = re.findall(r"([\u4e00-\u9fff]+)\s+([0-9|,|.]{4,})\s+([0-9|,|.]{4,})\s+([0-9|,|.]{4,})\s+([0-9|,|.]{4,})",company_list[i])
        # if len(lot_of_number_condition_list) != 0:
        #     lots_of_number_file.write(company_list[i])

        #找出所有的 字 数 数 的情况,加入list
        patten_list = re.findall(r"([\u4e00-\u9fff|:|：]+|合\s*计|销\s*售\s*费\s*用|[\u4e00-\u9fff]+\s\s?[\u4e00-\u9fff]+)\s+([0-9|,|.|\(|\)|\（|\）|．|，]{3,})\s+([0-9|,|.|\(|\)|\（|\）|．|，]+)",company_list[i])

        #找出所有的  字 数 字   的情况,其中  数字的值要大于五位数,可区分好多年份等单独数字
        only_one_condition_list = re.findall(r"([\u4e00-\u9fff|-]+)\s+([0-9|,|.]{5,})\s+(?=[\u4e00-\u9fff|-]+)",company_list[i])
        for only_one_condition in only_one_condition_list:
            add_item = (only_one_condition[0],only_one_condition[1],-1)
            # print(add_item)
            patten_list.append(add_item)

        # if(len(only_one_condition_list) != 0):
        #     print(code)
        #     print(only_one_condition_list)

    #   仅含有最终的一个销售费用
        if len(patten_list) == 0:
            qiguai_condition_list = re.findall(r"(销售费用|销售费用（元）|销售费用（万元）|销\s*售\s*费\s*用)\s+(\S)+\s+([0-9|,|.|\(|\)]+)\s+([0-9|,|.|\(|\)]+)",company_list[i])
            if len(qiguai_condition_list) != 0:
                qiguai_item = (qiguai_condition_list[0][2],qiguai_condition_list[0][3],-2)
                patten_list.append(qiguai_item)
        if len(patten_list) == 0:
            qiguai_condition_list = re.findall(r"(销售费用|销售费用（元）)\s+([0-9|,|.|\(|\)]{4,})",company_list[i])
            if len(qiguai_condition_list) != 0:
                qiguai_item = (qiguai_condition_list[0][0],qiguai_condition_list[0][1],-1)
                patten_list.append(qiguai_item)

        if len(patten_list) == 0:
            no_double_num_file.write(company_list[i])
            continue

        #缺少合计的时候,找到其他,其他后三,插入合计
        no_has_all_condition = re.findall(r"(其他)\s+([0-9|,|.]{4,})\s+([0-9|,|.]{4,})\s+([0-9|,|.]{4,})\s+([0-9|,|.]{4,})",company_list[i])
        patten_num = len(patten_list)
        if len(no_has_all_condition) != 0:
            all_condition= ('合计',no_has_all_condition[0][3],no_has_all_condition[0][4])
            patten_list.append(all_condition)
            # no_all_file.write(company_list[i])

        for patten in patten_list:
            # print(final_excel_row)
            final_excel_row+=1
            if patten[2] == -1:
                other = "不确定是本期还是上期"
                row_now = [code, short_name, count_time, -1, patten[0], patten[1], money, other, company_name]
                for i in range(len(biaotou)):
                    # print(i)
                    final_excel_sheet.cell(row=final_excel_row, column=i + 1).value = row_now[i]
            elif patten[2] == -2:
                other = "仅含有销售费用的整体部分"
                row_now = [code, short_name, count_time, 1,"整体销售费用", patten[0], money, other, company_name]
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i + 1).value = row_now[i]
                final_excel_row += 1
                row_now = [code, short_name, count_time, 0, "整体销售费用", patten[1], money, other, company_name]
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i + 1).value = row_now[i]
            else:

                row_now = [code,short_name,count_time,1,patten[0],patten[1],money,other,company_name]
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
                final_excel_row+=1
                row_now = [code, short_name, count_time, 0, patten[0], patten[2],money, other, company_name]
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i+1).value = row_now[i]
    print(txtname)
    print(final_excel_row)
final_excel.save(filedestination + final_file + ".xlsx")

