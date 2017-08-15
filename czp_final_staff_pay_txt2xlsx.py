# encoding:utf-8
import xlwt
import re
from xlsxwriter import *
import openpyxl
import glob

filelocation = '/Users/chenzhaopeng/PycharmProjects/final_txt2xlsx/final_staff_pay_data/'
read_file_form = "txt"
fileform = "xlsx"
filedestination = "/Users/chenzhaopeng/PycharmProjects/final_txt2xlsx/"
final_file = "merge_all_finance_information"


# 读取txt文件夹里有多少文件
filearray = []
for filename in glob.glob(filelocation + "*." + read_file_form):
    filearray.append(filename)
    print(filename)

print("在文件夹内有 %d 个txt文档" % len(filearray))

# 添加表头
biaotou = ['股票代码', '股票简称', '会计期间', '项目名称', '上期项目金额', '本期增加额', '本期减少额', '本期项目金额', '计价货币', '说明', '公司名称']
final_excel = openpyxl.Workbook()
final_excel_sheet = final_excel.active
final_excel_sheet.append(biaotou)

final_excel_row = 1

no_all_pay_table = 0
season_no_pay_table = 0
no_short_pay_table = 0
no_plan_pay_table = 0
no_patten_number = 0
no_patten_number_file = open("no_patten_number_file.txt","w+")
no_all_pay_table_file = open("no_all_pay_table_file.txt",'w+')
no_short_pay_table_file = open("no_short_pay_table_file.txt",'w+')
no_plan_pay_table_file = open("no_plan_pay_table_file.txt",'w+')
season_no_pay_number_file = open('season_no_pay_number_file.txt','w+')
is_season_file = None

#找到第一个表的所有 字 数 数 数,返回 patten_list
def find_the_first_table(the_all_line):
    patten_list = ''
    #先找出所有 字 数 数 数  的情况,   1    一  、           合     计                                               数字 -
    patten_list = re.findall(r'([0-9]?[一二三四五六七]?[、]?[其中：]?[\u4e00-\u9fff|、|-]+[\s]{0,2}[\u4e00-\u9fff]{0,3})\s+'
                             r'([-]?[0-9|.|,|)|(]{3,}[.]?[0-9]{,2}[\s]{0,2}[1,9]?|0|-|--|—|0.00)\s+'
                             r'([-]?[0-9|.|,|)|(]{3,}[.]?[0-9]{,2}[\s]{0,2}[1,9]?|0|-|--|—|0.00)\s+'
                             r'([-]?[0-9|.|,|)|(]{3,}[.]?[0-9]{,2}[\s]{0,2}[1,9]?|0|-|--|—|0.00)\s+'
                             r'([-]?[0-9|.|,|)|(]{3,}[.]?[0-9]{,2}[\s]{0,2}[1,9]?|0|-|--|—|0.00)',the_all_line)
    #存在个别   字 数  数的情况
    double_patten_list = re.findall(r'([0-9]?[一二三四五六七]?[、]?[其中：]?[\u4e00-\u9fff|、|-]+[\s]{0,2}[\u4e00-\u9fff]{0,3})\s+'
                                    r'([-]?[0-9|.|,|)|(]{3,}[.]?[0-9]?[\s]{0,2}[1,9]?|0|-|--|—|0.00)\s+'
                                    r'([-]?[0-9|.|,|)|(]{3,}[.]?[0-9]?[\s]{0,2}[1,9]?|0|-|--|—|0.00)\s+[0?]\s+'
                                    r'?([0-9]?[一二三四五六七]?[、]?[其中：]?[\u4e00-\u9fff|、|-]+[\s]{0,2}[\u4e00-\u9fff]{0,3})\s+',the_all_line)
    for double_patten in double_patten_list:
        double_patten = [double_patten[0],0,double_patten[1],double_patten[2],0]
        patten_list.append(double_patten)
    return patten_list

#把相应的项目转为需求的项目
def match_item(patten,a):
    # print(len(patten))
    if len(re.findall(r"短\s*期\s*薪\s*酬",patten[0])):
        string_now = "短期薪酬合计"
    if len(re.findall(r"设\s*定\s*提\s*存\s*计\s*划",patten[0])):
        string_now = "离职后福利-设定提存计划合计"
    if len(re.findall(r"工资|奖金|津贴|补贴",patten[0])):
        string_now = "短期薪酬,其中:1.工资、奖金、津贴和补贴"
    elif len(re.findall(r"职\s*工\s*福\s*利\s*费",patten[0])):
        string_now = "短期薪酬,其中:2.职工福利费"
    elif len(re.findall(r"社\s*会\s*保\s*险\s*费",patten[0])):
        string_now = "短期薪酬,其中:3.社会保险费"
    elif len(re.findall(r"医\s*疗\s*保\s*险\s*费", patten[0])):
        string_now = "短期薪酬,其中:3.社会保险费,其中:医疗保险费"
    elif len(re.findall(r"工\s*伤\s*保\s*险\s*费", patten[0])):
        string_now = "短期薪酬,其中:3.社会保险费,其中:工伤保险费"
    elif len(re.findall(r"生\s*育\s*保\s*险\s*费", patten[0])):
        string_now = "短期薪酬,其中:3.社会保险费,其中:生育保险费"
    elif len(re.findall(r"伤\s*残\s*就\s*业\s*金", patten[0])):
        string_now = "短期薪酬,其中:3.社会保险费,其中:伤残就业金"
    elif len(re.findall(r"住\s*房\s*公\s*积\s*金", patten[0])):
        string_now = "短期薪酬,其中:4.住房公积金"
    elif len(re.findall(r"住\s*房\s*公\s*积\s*金", patten[0])):
        string_now = "短期薪酬,其中:4.住房公积金"
    elif len(re.findall(r"工\s*会\s*经\s*费|职\s*工\s*教\s*育", patten[0])):
        string_now = "短期薪酬,其中:5.工会经费和职工教育费"
    elif len(re.findall(r"短\s*期\s*带\s*薪", patten[0])):
        string_now = "短期薪酬,其中:6.短期带薪缺勤"
    elif len(re.findall(r"短\s*期\s*利\s*润", patten[0])):
        string_now = "短期薪酬,其中:7.短期利润分享计划"
    elif len(re.findall(r"其\s*他", patten[0])):
        if a == 0:
            string_now = '一年内到期的其他福利'
        elif a == 1:
            string_now = "短期薪酬,其中:8.其他"
        elif a == 2:
            string_now = '离职后福利-设定提存计划,其中:4.其他'
    elif len(re.findall(r"养\s*老\s*保\s*险", patten[0])):
        string_now = '离职后福利-设定提存计划,其中:1.基本养老保险'
    elif len(re.findall(r"失\s*业\s*保\s*险", patten[0])):
        string_now = '离职后福利-设定提存计划,其中:2.失业保险'
    elif len(re.findall(r"企\s*业\s*年\s*金", patten[0])):
        string_now = '离职后福利-设定提存计划,其中:3.企业年金缴费'
    elif len(re.findall(r"辞\s*退\s*福\s*利", patten[0])):
        string_now = '辞退福利'
    elif len(re.findall(r"合\s*计", patten[0])):
        if a == 0:
            string_now = '总计数'
        elif a == 1:
            string_now = "短期薪酬,总计数"
        elif a == 2:
            string_now = '离职后福利-设定提存计划,总计数'
    else:
        string_now = patten[0]
    patten = [string_now,patten[1],patten[2],patten[3],patten[4]]
    return patten

import time
# 逐个读取文件
# for txtname in filearray:
for i in range(len(filearray)):
    start = time.clock()
    txtname = filearray[i]
    is_season_file = None
    if len(re.findall(r'season',txtname)) > 0:
        print("这是季度报")
        is_season_file = 1
    else:
        print("年报")

    txtfile = open(txtname, 'r')
    line = txtfile.read()

    # 先按照——————划分出不同公司
    company_list = line.split('---------------------------------')
    # print(company)
    # for i in range(100):
    for i in range(len(company_list) -1):
        # for company in company_list:
        data_time = None
        pro_name = None
        pro_num = None
        company_name = None
        current_time_increase = None
        current_time_loss = None
        other = None
        money = 'CNY'

        # each_line 公司里面的每个项目,空格分开
        each_line = company_list[i].split('\n')
        # print(len(each_line))
        # print(each_line)
        if len(each_line) < 6:
            continue
        code = each_line[1]
        short_name = each_line[2]
        count_time = each_line[3]
        company_name = each_line[4]
        # print(company_name)
        the_first_line = each_line[5]
        the_first_line = the_first_line.replace('','')
        the_first_line = the_first_line.replace('','')
        # print(the_first_line)
        #07-16年数据
        # print(is_season_file)
        if is_season_file:
            other = "季度报表,仅含有应付职工薪酬的期初和期末余额"
            # each_item = each_line[4].split()
            #   字   数    数     数,    取第一和第三个
            # print(each_line[4])
            if len(re.findall(r'(应\s*付\s*职\s*工\s*薪\s*酬)\s*[\S]{,4}\s*([0-9|,|.|-]{5,}|0.00)\s*([0-9|,|.|-]{5,}|0.00)\s*([0-9|,|.|-]{5,}|0.00)\s*([0-9|,|.|-]{5,}|0.00)',the_first_line)) > 0:
                #写数据
                patten = re.findall(r'(应\s*付\s*职\s*工\s*薪\s*酬)\s*[\S]{,4}\s*([0-9|,|.|-]{5,}|0.00)\s*([0-9|,|.|-]{5,}|0.00)\s*([0-9|,|.|-]{5,}|0.00)\s*([0-9|,|.|-]{5,}|0.00)',the_first_line)
                # print("薪酬加四个数")
                final_excel_row += 1
                #biaotou = ['股票代码', '股票简称', '会计期间', '项目名称', '上期项目金额', '本期增加额', '本期减少额', '本期项目金额', '计价货币', '说明', '公司名称']
                # print(patten)
                #留意,是左边是期末,右边是期初
                row_now = [code, short_name, count_time, patten[0][0], patten[0][3],'None' ,'None',patten[0][1], money,other, company_name]
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i + 1).value = row_now[i]
            elif len(re.findall(r'(应\s*付\s*职\s*工\s*薪\s*酬)\s*([0-9|,|.|-]{5,})\s*([0-9|,|.|-]{5,})',the_first_line)) > 0:
                #写数据
                patten = re.findall(r'(应\s*付\s*职\s*工\s*薪\s*酬)\s*([0-9|,|.|-]{5,})\s*([0-9|,|.|-]{5,})',the_first_line)

                final_excel_row += 1
                #只有两个的话
                row_now = [code, short_name, count_time, patten[0][0], patten[0][2],'None' ,'None',patten[0][1], money,
                           other, company_name]
                for i in range(len(biaotou)):
                    final_excel_sheet.cell(row=final_excel_row, column=i + 1).value = row_now[i]
            else:
                season_no_pay_number_file.write(company_list[i])
                season_no_pay_table += 1

        else:
            # 07-14年的数据只有五行
            # print("这是07-16年 年报,半年报")
            if len(each_line) <= 7:
                # print("这是07-14年 年报,半年报")
                if the_first_line == '没有应付职工薪酬表格':
                    no_all_pay_table_file.write(company_list[i])
                    no_all_pay_table += 1
                    continue
                else:
                    patten_list = find_the_first_table(the_first_line)
                    # other = '仅含有一个总的表格'
                    for patten in patten_list:
                        # print(patten[0])
                        patten = match_item(patten,0)
                        # print(patten[0])
                        #开始写入数据
                        final_excel_row+=1
                        row_now = [code,short_name,count_time,patten[0],patten[1],patten[2],patten[3],patten[4],money,other,company_name]
                        for i in range(len(biaotou)):
                            final_excel_sheet.cell(row=final_excel_row, column=i + 1).value = row_now[i]
            #15-16年的数据有三个子表
            else:
                # print("这是15-16的年报半年报")
                if each_line[5] == '没有应付职工薪酬列示的表格':
                    no_all_pay_table_file.write(company_list[i])
                    no_all_pay_table+= 1
                else:
                    patten_list = find_the_first_table(each_line[5])
                    # other = '仅含有一个总的表格'
                    if len(patten_list) == 0 :
                        no_patten_number_file.write(company_list[i])
                        no_patten_number+=1
                    for patten in patten_list:
                        patten = match_item(patten, 0)

                        # 开始写入数据
                        final_excel_row += 1
                        row_now = [code, short_name, count_time, patten[0], patten[1], patten[2], patten[3], patten[4],
                                   money, other, company_name]
                        for i in range(len(biaotou)):
                            final_excel_sheet.cell(row=final_excel_row, column=i + 1).value = row_now[i]
                if each_line[6] == '没有短期薪酬列示的表格':
                    no_short_pay_table_file.write(company_list[i])
                    no_short_pay_table+= 1
                else:
                    patten_list = find_the_first_table(each_line[6])
                    if len(patten_list) == 0:
                        no_patten_number_file.write(company_list[i])
                        no_patten_number += 1
                    for patten in patten_list:
                        patten = match_item(patten, 1)

                        # 开始写入数据
                        final_excel_row += 1
                        row_now = [code, short_name, count_time, patten[0], patten[1], patten[2], patten[3],
                                   patten[4],
                                   money, other, company_name]
                        for i in range(len(biaotou)):
                            final_excel_sheet.cell(row=final_excel_row, column=i + 1).value = row_now[i]
                if each_line[7] == '没有提存计划列示的表格':
                    no_plan_pay_table_file.write(company_list[i])
                else:
                    patten_list = find_the_first_table(each_line[7])
                    if len(patten_list) == 0:
                        no_patten_number_file.write(company_list[i])
                        no_patten_number += 1
                    # other = '仅含有一个总的表格'
                    for patten in patten_list:
                        patten = match_item(patten, 2)

                        # 开始写入数据
                        final_excel_row += 1
                        row_now = [code, short_name, count_time, patten[0], patten[1], patten[2], patten[3],
                                   patten[4],
                                   money, other, company_name]
                        for i in range(len(biaotou)):
                            final_excel_sheet.cell(row=final_excel_row, column=i + 1).value = row_now[i]

    print(txtname)
    print(final_excel_row)
    final_excel.save(filedestination + final_file + ".xlsx")
    end = time.clock()
    print("花了%d 秒"% int(end-start))
print(no_patten_number)
print(no_short_pay_table)
print(no_plan_pay_table)
print(no_all_pay_table)